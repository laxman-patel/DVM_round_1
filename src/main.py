from openpyxl import load_workbook
from Classes import timetable, course, section, slot

def populate_course(filename, my_timetable):
    workbook = load_workbook(filename)
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        
        if i == 0:
            continue
        
        course_name = row[0]
        exam_date = row[1]
        prof_name = row[2]
        course_units = row[3]

        my_course = course.Course(course_name, exam_date, prof_name, course_units)
        my_timetable.enroll_course(my_course)


def populate_section(section_type,slot,my_course):
    my_section = section.Section(section_type,slot,my_course)
    my_course.populate_section(my_section)
    

def main():
    my_timetable = timetable.Timetable()

    while True:
        print("\n1. Populate courses from file")
        print("2. Display all courses")
        print("3. Add section to course")
        print("4. Check clashes")
        print("5. Export to CSV")
        print("6. Exit")

        choice = input("\nEnter your choice: ")

        if choice == '1':
            try:
                filename = input("Enter the filename: ")
                populate_course(filename, my_timetable)

                print("File imported successfully!")
            except FileNotFoundError:
                print("File not found. Please try again.")
                continue
       
        elif choice == '2':
            for course in my_timetable.courses:
                print(course)
       
        elif choice == '3':
            course_name = input("Enter the course name: ")
            section_type = input("Enter the section type: ")
            day = input("Enter the day: ")
            start_time = input("Enter the start time in HH:MM format: ")
            end_time = input("Enter the end time in HH:MM format: ")
      
            my_course = None
      
            for course in my_timetable.courses:
                if course.course_name == course_name:
                    my_course = course
                    break
           
            if my_course == None:
                print("Course not found")
                continue
          
            my_slot = slot.Slot(day, start_time, end_time)
            populate_section(section_type,my_slot,my_course)
      
        elif choice == '4':
            if my_timetable.check_clashes():
                print("There are clashes in your timetable")
            else:
                print("There are no clashes in your timetable")
        
        elif choice == '5':
            try:
                filename = input("Enter the filename: ")
                my_timetable.export_to_csv(filename)
            except FileNotFoundError:
                print('File not found. Please try again.')

        elif choice == '6':
            break
       
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()